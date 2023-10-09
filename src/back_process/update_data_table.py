import os

import openpyxl
import pandas as pd
import PyPDF2

from library.constants import *
import pandas


def update_cadence(_id: str) -> None:
    workbook = openpyxl.load_workbook(PATH_EXCEL)
    sheet = sheet = workbook[workbook.sheetnames[0]]

    value_subject = pd.read_excel(PATH_EXCEL, index_col=0).loc[_id]
    order_shoes = value_subject['order_nike':'order_individual_3'].sort_values(ascending=True)

    # get value from pdf report
    for index_shoe, shoe in enumerate(order_shoes.keys()):
        name = shoe.split('_')[-1]

        if name == 'ascis':
            name = 'asics'
        path_pdf = []
        for dirpath, subdirs, files in os.walk(os.path.join(PATH_PRESSURE_REPORT, _id)):
            path_pdf = os.path.join(dirpath, [file for file in files if name in file][0])
            pdf_file = open(path_pdf, 'rb')
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            for page in pdf_reader.pages:
                lines = page.extract_text().split('\n')

                for line in lines:

                    if 'Cadence'.lower() in line.lower():
                        value_cadence = line.split(' ')[2].split('±')[0]
                        value_subject[f'cadence_{index_shoe + 1}'] = value_cadence
            pdf_file.close()

    # Iterate over subject row excel data
    for row in sheet.iter_rows(min_row=2):  # Start from the second row (excluding headers)
        if row[0].value == _id:
            row_index = row[0].row
    try:
        for index_column, value in enumerate(value_subject):
            sheet.cell(row=row_index, column=index_column + 2).value = value

        workbook.save(os.path.join(PATH_DATA, 'subjects_running_economy.xlsx'))
    except:
        raise Exception(f"ID : {_id} not found")

    workbook.close()
